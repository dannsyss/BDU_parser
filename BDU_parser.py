from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# Структура для хранения данных об уязвимости
class Vulnerability:
    def __init__(self, id, description):
        self.id = id
        self.description = description

# Функция для парсинга HTML и извлечения идентификаторов уязвимостей
def parse_html(html_file):
    vulnerabilities = []
    with open(html_file, "r", encoding="utf-8") as file:
        soup = BeautifulSoup(file, "html.parser")
        # Находим все строки с идентификаторами уязвимостей
        for td in soup.find_all("td", class_="bdu"):
            vulnerabilities.append(td.text.strip())
    return vulnerabilities

# Функция для парсинга Excel и извлечения описаний уязвимостей
def parse_excel(excel_file, ids):
    vulnerabilities = []
    workbook = load_workbook(excel_file)
    worksheet = workbook.active

    # Читаем данные из Excel
    for row in worksheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
        vuln_id = row[0]  # Идентификатор уязвимости
        description = row[2]  # Описание уязвимости

        # Проверяем, есть ли идентификатор в списке
        if vuln_id in ids:
            vulnerabilities.append(Vulnerability(vuln_id, description))

    return vulnerabilities

# Функция для создания новой таблицы
def create_table(vulnerabilities, output_file):
    workbook = Workbook()
    worksheet = workbook.active

    # Заголовки таблицы
    worksheet.append(["№", "BDU", "Наименование уязвимости"])

    # Заполняем таблицу данными
    for index, vuln in enumerate(vulnerabilities, start=1):
        worksheet.append([index, vuln.id, vuln.description])

    # Сохраняем файл
    workbook.save(output_file)

# Основная функция
def main():
    html_file = "ScanOval_Report_26_02_2025_bd6cc06f-cc70-402d-96ed-1c711849a567.html"
    excel_file = "vullist.xlsx"
    output_file = "output.xlsx"

    # Парсим HTML и извлекаем идентификаторы уязвимостей
    ids = parse_html(html_file)

    # Парсим Excel и извлекаем описания уязвимостей
    vulnerabilities = parse_excel(excel_file, ids)

    # Создаем новую таблицу
    create_table(vulnerabilities, output_file)

    print(f"Таблица успешно создана: {output_file}")

if __name__ == "__main__":
    main()