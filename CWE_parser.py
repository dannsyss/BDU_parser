import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


def parse_excel(excel_file):
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    data = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Разделяем типы ошибок CWE по запятой и убираем лишние пробелы
        cwe_types = [cwe.strip() for cwe in str(row[-1]).split(',')]

        # Для каждого типа ошибки CWE создаём отдельную запись
        for cwe in cwe_types:
            data.append({
                'Идентификатор': row[0],  # Идентификатор уязвимости
                'Наименование уязвимости': row[1],  # Наименование уязвимости
                'Тип ошибки CWE': cwe  # Тип ошибки CWE
            })

    return pd.DataFrame(data)


# Кэш для хранения CAPEC и их уровней угроз
capec_cache = {}


# Функция для получения CAPEC по CWE
def get_capec(cwe_id):
    if cwe_id in capec_cache:
        print(f"Используем кэш для CWE: {cwe_id}")
        return capec_cache[cwe_id]

    url = f"https://cwe.mitre.org/data/definitions/{cwe_id}.html"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    capec_table = soup.find('div', id='Related_Attack_Patterns')
    if capec_table:
        rows = capec_table.find_all('tr')[1:]  # Пропускаем заголовок
        capecs = []
        for row in rows:
            cols = row.find_all('td')
            capec_id = cols[0].text.strip()
            attack_name = cols[1].text.strip()
            capecs.append(capec_id)
        capec_cache[cwe_id] = capecs  # Сохраняем в кэш
        return capecs
    return []


# Функция для получения Likelihood Of Attack по CAPEC
def get_likelihood(capec_id):
    if capec_id in capec_cache:
        print(f"Используем кэш для CAPEC: {capec_id}")
        return capec_cache[capec_id]

    url = f"https://capec.mitre.org/data/definitions/{capec_id.split('-')[1]}.html"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    likelihood_div = soup.find('div', id='Likelihood_Of_Attack')
    if likelihood_div:
        likelihood = likelihood_div.find('div', class_='detail').text.strip()
        capec_cache[capec_id] = likelihood  # Сохраняем в кэш
        return likelihood
    return "No chance"


# Основной скрипт
def main():
    # Чтение данных из Excel файлов
    print("Чтение данных из vullist.xlsx...")
    vullist_df = parse_excel('vullist.xlsx')
    print("Чтение данных из output.xlsx...")
    output_df = pd.read_excel('output.xlsx')

    # Создание новой таблицы для результатов
    results = []

    # Общее количество строк для обработки
    total_rows = len(output_df)
    print(f"Всего строк для обработки: {total_rows}")

    # Обработка каждой строки в output.xlsx
    for index, row in output_df.iterrows():
        bdu = row['BDU']
        vuln_name = row['Наименование уязвимости']

        # Вывод прогресса
        print(f"Обработка строки {index + 1} из {total_rows}...")

        # Поиск соответствующей строки в vullist_df
        cwe_rows = vullist_df[vullist_df['Идентификатор'] == bdu]
        if not cwe_rows.empty:
            for _, cwe_row in cwe_rows.iterrows():
                cwe_id = cwe_row['Тип ошибки CWE']
                if pd.notna(cwe_id):  # Проверяем, что CWE не пустое
                    print(f"Найден CWE: {cwe_id}")
                    cwe_number = cwe_id.split('-')[1]  # Извлекаем номер CWE

                    # Получаем CAPEC для данного CWE
                    capecs = get_capec(cwe_number)
                    print(f"Найдено CAPEC: {len(capecs)}")

                    # Группируем CAPEC по Likelihood Of Attack
                    high_capecs = []
                    medium_capecs = []
                    low_capecs = []
                    no_chance_capecs = []

                    for capec in capecs:
                        likelihood = get_likelihood(capec)
                        if likelihood == "High":
                            high_capecs.append(capec)
                        elif likelihood == "Medium":
                            medium_capecs.append(capec)
                        elif likelihood == "Low":
                            low_capecs.append(capec)
                        else:
                            no_chance_capecs.append(capec)

                    # Добавляем результаты в таблицу
                    results.append({
                        '№': index + 1,
                        'BDU': bdu,
                        'CWE': cwe_id,
                        'CAPEC High': ', '.join(high_capecs),
                        'CAPEC Medium': ', '.join(medium_capecs),
                        'CAPEC Low': ', '.join(low_capecs),
                        'No chance': ', '.join(no_chance_capecs)
                    })
                else:
                    print(f"CWE не найден для BDU: {bdu}")
        else:
            print(f"BDU {bdu} не найден в vullist.xlsx")

    # Сохранение результатов в новый Excel файл
    print("Сохранение результатов в threat_level.xlsx...")
    results_df = pd.DataFrame(results)
    results_df.to_excel('threat_level.xlsx', index=False)
    print("Результаты успешно сохранены в файл 'threat_level.xlsx'")


if __name__ == "__main__":
    main()