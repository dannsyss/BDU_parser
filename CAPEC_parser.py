import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


def parse_excel(excel_file):
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    data = []

    # Чтение данных из Excel, начиная со второй строки (первая строка — заголовки)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        data.append({
            'Идентификатор': row[0],  # Идентификатор уязвимости
            'Наименование уязвимости': row[1],  # Наименование уязвимости
            'Тип ошибки CWE': row[-1]  # Тип ошибки CWE (последняя колонка)
        })

    return pd.DataFrame(data)


# Функция для получения CAPEC по CWE
def get_capec(cwe_id):
    url = f"https://cwe.mitre.org/data/definitions/{cwe_id}.html"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    capec_table = soup.find('div', id='Related_Attack_Patterns')
    if capec_table:
        rows = capec_table.find_all('tr')[1:]  # Пропускаем заголовок
        capecs = []
        for row in rows:
            cols = row.find_all('td')
            capec_id = cols[0].text.strip()
            attack_name = cols[1].text.strip()
            capecs.append((capec_id, attack_name))
        return capecs
    return []


# Основной скрипт
def main():
    # Чтение данных из Excel файлов
    vullist_df = parse_excel('vullist.xlsx')
    output_df = pd.read_excel('output.xlsx')

    # Создание новой таблицы для результатов
    results = []
    unique_capecs = set()  # Для отслеживания уникальных CAPEC
    row_number = 1  # Нумерация строк

    # Обработка каждой строки в output.xlsx
    for index, row in output_df.iterrows():
        bdu = row['BDU']
        vuln_name = row['Наименование уязвимости']

        # Поиск соответствующей строки в vullist_df
        cwe_row = vullist_df[vullist_df['Идентификатор'] == bdu]
        if not cwe_row.empty:
            cwe_id = cwe_row.iloc[0]['Тип ошибки CWE']
            if pd.notna(cwe_id):  # Проверяем, что CWE не пустое
                capecs = get_capec(cwe_id.split('-')[1])  # Извлекаем номер CWE
                for capec in capecs:
                    if capec[0] not in unique_capecs:  # Проверяем, что CAPEC уникален
                        unique_capecs.add(capec[0])
                        results.append({
                            '№': row_number,
                            'CAPEC': capec[0],
                            'Наименование атаки': capec[1]
                        })
                        row_number += 1  # Увеличиваем номер строки

    # Сохранение результатов в новый Excel файл
    results_df = pd.DataFrame(results)
    results_df.to_excel('capec_results.xlsx', index=False)
    print("Результаты сохранены в файл 'capec_results.xlsx'")


if __name__ == "__main__":
    main()